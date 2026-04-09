import openpyxl
import requests

def get_cid(name):
    url = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/name/{name}/cids/JSON"
    response = requests.get(url)
    if response.status_code == 200:
        return response.json()["IdentifierList"]["CID"][0]
    else:
        return None

wb = openpyxl.load_workbook("metabolites_completed.xlsx")
ws = wb.active

#print(ws.title)
#print(ws.max_row)
#print(ws.max_column)

for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True)):
    name = row[1]
    cid = get_cid(name)
    ws.cell(row=i+2, column=5).value = cid

wb.save("metabolites_step1_PubChem.xlsx")
print("Saved successfully!")
