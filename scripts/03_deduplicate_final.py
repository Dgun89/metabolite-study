"""
03_deduplicate_final.py

human_final.xlsx / mouse_final.xlsx 최신본에 남아 있는 중복 화합물 행을 제거한다.

중복 정의:
    같은 화합물이 여러 행으로 반복된 것 = "Database ID"(COCONUT CNP id)가 동일한 행.
    최신 데이터에서는 이런 행이 나머지 컬럼까지 완전히 동일(fully-identical)하므로,
    Database ID 기준 첫 행만 남기고(keep="first") 제거하면 완전 중복 제거와 일치한다.

주의:
    InChIKey가 같아도 Database ID가 다르면 서로 다른 입체이성질체일 수 있다
    (예: DL-ornithine=CNP0102727.0 vs L-ornithine=CNP0102727.1).
    이런 경우는 제거하지 않고 참고용으로 콘솔에만 보고한다.

출력:
    3-sheet(Sheet1/Legend/Summary) 구조와 서식(헤더 색 등)을 그대로 유지한 채
    Sheet1의 중복 행만 삭제하고 Summary의 카운트/커버리지 셀을 재계산해 갱신한다.
    원본은 건드리지 않고 {species}_final_dedup.xlsx 로 저장한다.

사용법:
    python 03_deduplicate_final.py                     # data/{species}/final/*_final.xlsx 자동 검색
    python 03_deduplicate_final.py --human path.xlsx --mouse path.xlsx
    python 03_deduplicate_final.py --species mouse     # 한 종만
"""
import argparse
import datetime as dt
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

KEY = "Database ID"          # 중복 판정 기준 컬럼
DATA_SHEET = "Sheet1"
SUMMARY_SHEET = "Summary"

CANDIDATE_DIRS = {
    "human": ["data/human/final", "final", "."],
    "mouse": ["data/mouse/final", "final", "."],
}

# Summary 재계산에 쓰는 매핑
DB_ID_COLS = {"PubChem": "PubChem", "KEGG": "KEGG", "HMDB": "HMDB", "ChEBI": "ChEBI"}
ENZYME_COLS = {
    "KEGG": "kegg_enzymes",
    "HMDB": "hmdb_enzymes",
    "Reactome": "reactome_catalysts",
    "BRENDA": "brenda_enzymes",
}


def find_file(species: str, given: str | None) -> str:
    if given:
        fp = Path(given)
        if fp.exists():
            return str(fp)
        raise FileNotFoundError(f"{species}: {given} 을(를) 찾을 수 없습니다.")
    for d in CANDIDATE_DIRS.get(species, []):
        # 이미 dedup된 결과물은 제외하고 원본만 검색
        matches = sorted(
            m for m in Path(d).glob(f"*{species}_final.xlsx")
            if "_dedup" not in m.name
        )
        if matches:
            if len(matches) > 1:
                print(f"[안내] {species}: {d} 에서 후보 {len(matches)}개 중 "
                      f"{matches[0].name} 사용")
            return str(matches[0])
    raise FileNotFoundError(
        f"{species}_final.xlsx 를 못 찾았습니다. --{species} 로 경로를 지정하세요."
    )


def analyze(df: pd.DataFrame):
    """중복 현황을 조사해 (제거대상 행번호 리스트, 리포트 dict) 반환."""
    n = len(df)
    dup_mask = df.duplicated(subset=[KEY], keep="first")   # 2번째 이후 발생 = 제거대상
    n_remove = int(dup_mask.sum())
    n_unique = int(df[KEY].nunique(dropna=True))

    # InChIKey는 같은데 Database ID가 다른(=서로 다른 화합물로 취급되는) 경우 보고
    ik_note = []
    if "InChIKey" in df.columns:
        dedup_ids = df[~dup_mask]
        vc = (dedup_ids.dropna(subset=["InChIKey"])
                       .groupby("InChIKey")[KEY].nunique())
        collide = vc[vc > 1]
        for ik in collide.index:
            ids = sorted(dedup_ids[dedup_ids["InChIKey"] == ik][KEY].unique())
            names = dedup_ids[dedup_ids["InChIKey"] == ik]["compound_name"].unique()
            ik_note.append((ik, ids, list(names)))

    report = {
        "rows_before": n,
        "rows_removed": n_remove,
        "rows_after": n - n_remove,
        "unique_ids": n_unique,
        "inchikey_collisions": ik_note,
    }
    # 제거대상의 0-based 위치 → 엑셀 행번호(헤더가 1행이므로 +2)
    remove_positions = [i for i, v in enumerate(dup_mask.tolist()) if v]
    return remove_positions, report


def recompute_summary(dedup: pd.DataFrame) -> dict:
    """dedup된 Sheet1로부터 Summary의 Item→(count, coverage) 값을 재계산."""
    total = len(dedup)
    out = {}

    def put(item, count):
        cov = f"{count / total * 100:.1f}%" if total else "0%"
        out[item] = (count, "100%" if item == "Total compounds" else cov)

    put("Total compounds", total)
    for item in ("InChIKey", "SMILES"):
        if item in dedup.columns:
            put(item, int(dedup[item].notna().sum()))
    for item, col in DB_ID_COLS.items():
        if col in dedup.columns:
            put(item, int(dedup[col].notna().sum()))
    if "classification" in dedup.columns:
        cls = dedup["classification"].astype(str).str.strip().str.lower()
        put("Endogenous", int((cls == "endogenous").sum()))
        put("Exogenous", int((cls == "exogenous").sum()))
        put("Unverified", int((cls == "unverified").sum()))

    # 효소 소스 Venn (4개 소스의 존재 조합)
    present = {}
    for src, col in ENZYME_COLS.items():
        present[src] = dedup[col].notna() if col in dedup.columns else pd.Series(False, index=dedup.index)
    order = ["KEGG", "HMDB", "Reactome", "BRENDA"]
    sig = dedup.index.to_series().map(
        lambda i: tuple(s for s in order if present[s].loc[i])
    )

    def label(combo):
        if len(combo) == 1:
            return f"{combo[0]} only"
        if len(combo) == 4:
            return "All 4 sources"
        return " + ".join(combo)

    counts = sig[sig.map(len) > 0].map(label).value_counts()
    for lab, c in counts.items():
        put(lab, int(c))
    any_enz = int(sig.map(len).gt(0).sum())
    put("Total (unique)", any_enz)
    return out


def write_dedup(src_path: str, remove_positions: list[int],
                summary_vals: dict, out_path: str):
    """openpyxl로 원본을 열어 서식 유지한 채 중복 행 삭제 + Summary 갱신."""
    wb = load_workbook(src_path)
    ws = wb[DATA_SHEET]
    # 엑셀 행번호로 변환(헤더 1행), 아래에서 위로 삭제해야 인덱스가 안 밀림
    excel_rows = sorted((p + 2 for p in remove_positions), reverse=True)
    for r in excel_rows:
        ws.delete_rows(r, 1)

    # Summary 갱신: B열=Item, C열=Count, D열=Coverage
    if SUMMARY_SHEET in wb.sheetnames:
        sm = wb[SUMMARY_SHEET]
        for row in sm.iter_rows():
            item_cell = row[1] if len(row) > 1 else None
            if item_cell is None:
                continue
            item = str(item_cell.value).strip() if item_cell.value is not None else ""
            if item in summary_vals:
                count, cov = summary_vals[item]
                row[2].value = count
                row[3].value = cov
            # 'Last updated: ...' 갱신
            if isinstance(item_cell.value, str) and item_cell.value.startswith("Last updated"):
                item_cell.value = f"Last updated: {dt.date.today().isoformat()}"
        # A열에 있을 수도 있는 Last updated 처리
        for row in sm.iter_rows(min_col=1, max_col=1):
            c = row[0]
            if isinstance(c.value, str) and c.value.startswith("Last updated"):
                c.value = f"Last updated: {dt.date.today().isoformat()} (deduplicated)"
    wb.save(out_path)


def process(species: str, path: str):
    df = pd.read_excel(path, sheet_name=DATA_SHEET)
    if KEY not in df.columns:
        raise KeyError(f"{path}: '{KEY}' 컬럼이 없습니다.")
    remove_positions, rep = analyze(df)
    dedup = df.drop(index=[df.index[p] for p in remove_positions]).reset_index(drop=True)
    summary_vals = recompute_summary(dedup)

    out_path = str(Path(path).with_name(f"{species}_final_dedup.xlsx"))
    write_dedup(path, remove_positions, summary_vals, out_path)

    print(f"\n===== {species} =====")
    print(f"입력: {path}")
    print(f"중복 제거 전 {rep['rows_before']}행 → 후 {rep['rows_after']}행 "
          f"(제거 {rep['rows_removed']}행, 고유 Database ID {rep['unique_ids']}개)")
    if rep["inchikey_collisions"]:
        print(f"[참고] InChIKey는 같지만 Database ID가 달라 유지된 화합물 "
              f"{len(rep['inchikey_collisions'])}건 (입체이성질체 등, 제거 안 함):")
        for ik, ids, names in rep["inchikey_collisions"][:10]:
            print(f"    {ik}: {', '.join(ids)}  ({' / '.join(str(n) for n in names)})")
        if len(rep["inchikey_collisions"]) > 10:
            print(f"    ... 외 {len(rep['inchikey_collisions']) - 10}건")
    print(f"저장: {out_path}  (3-sheet 유지, Summary 카운트 재계산)")
    return rep


def main():
    ap = argparse.ArgumentParser(description="final xlsx 중복 화합물 제거 (Database ID 기준)")
    ap.add_argument("--human", default=None)
    ap.add_argument("--mouse", default=None)
    ap.add_argument("--species", choices=["human", "mouse"], default=None,
                    help="한 종만 처리 (기본: 둘 다)")
    args = ap.parse_args()

    targets = [args.species] if args.species else ["human", "mouse"]
    for sp in targets:
        given = getattr(args, sp)
        path = find_file(sp, given)
        process(sp, path)


if __name__ == "__main__":
    main()
