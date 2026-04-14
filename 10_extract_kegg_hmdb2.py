# 10_extract_kegg_hmdb2.py
# Target: rows with PubChem CID but missing KEGG or HMDB in metabolites_step8.xlsx

import openpyxl
import requests
import time

def get_hmdb_kegg(cid):
    url = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug_view/data/compound/{cid}/JSON"
    response = requests.get(url)

    if response.status_code != 200:
        return {"hmdb": None, "kegg": None}

    data = response.json()
    result = {"hmdb": None, "kegg": None}

    sections = data["Record"]["Section"]
    for section in sections:
        if section.get("TOCHeading") == "Names and Identifiers":
            for subsection in section.get("Section", []):
                if subsection.get("TOCHeading") == "Other Identifiers":
                    for subsubsection in subsection.get("Section", []):
                        heading = subsubsection.get("TOCHeading")
                        if heading in ["HMDB ID", "KEGG ID"]:
                            for info in subsubsection.get("Information", []):
                                for val in info.get("Value", {}).get("StringWithMarkup", []):
                                    text = val.get("String")
                                    if heading == "HMDB ID":
                                        result["hmdb"] = text
                                    else:
                                        result["kegg"] = text
    return result

# ─────────────────────────────
# 1. 엑셀 읽기
# ─────────────────────────────
wb = openpyxl.load_workbook("metabolites_step8.xlsx")
ws = wb.active

# 컬럼 구조
# A(0): Database ID, B(1): InChIKey, C(2): QualitativeResults
# D(3): KEGG, E(4): HMDB, F(5): PubChem, G(6): reference

# ─────────────────────────────
# 2. 총 처리 대상 수 계산
# ─────────────────────────────
total = sum(1 for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True)
            if row[5] is not None and not (row[3] and row[4]))
print(f"Target: {total} rows")

# ─────────────────────────────
# 3. 루프 실행
# ─────────────────────────────
count = 0
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=0):
    cid = row[5].value    # .value 추가
    kegg = row[3].value   # .value 추가
    hmdb = row[4].value   # .value 추가

    if cid is None:       # CID 없으면 건너뜀
        continue
    if kegg and hmdb:     # 둘 다 있으면 건너뜀
        continue

    count += 1
    ids = get_hmdb_kegg(str(int(float(cid))))
    ws.cell(row=i+2, column=4).value = ids["kegg"] or kegg
    ws.cell(row=i+2, column=5).value = ids["hmdb"] or hmdb
    print(f"[{count}/{total}] {row[2]} → KEGG: {ids['kegg']} HMDB: {ids['hmdb']}")
    time.sleep(0.3)

# ─────────────────────────────
# 4. 저장
# ─────────────────────────────
wb.save("metabolites_step10_ver2.xlsx")
print("Saved successfully!")