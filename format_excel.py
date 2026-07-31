from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date

GROUPS = {
    "Basic Identifiers": {
        "columns": ["InChIKey", "compound_name", "SMILES"],
        "color": "BDD7EE",
        "description": "Core compound identifiers (primary key = InChIKey)"
    },
    "Dataset Membership": {
        "columns": ["datasets"],
        "color": "D6DCE4",
        "description": "Which source datasets this compound appears in (combined view only)"
    },
    "External DB IDs": {
        "columns": ["COCONUT", "PubChem", "KEGG", "HMDB", "ChEBI", "CAS", "DrugBank", "DrugCentral", "FooDB", "LIPID MAPS"],
        "color": "C6EFCE",
        "description": "Cross-reference IDs from external databases (UniChem-resolved)"
    },
    "Drug / Food": {
        "columns": ["drug_food"],
        "color": "FFF2CC",
        "description": "External drug/food signals (DrugBank/DrugCentral=drug, FooDB=food). Display-only flag, not a filter — rows are never dropped"
    },
    "Classification": {
        "columns": ["classification"],
        "color": "FFEB9C",
        "description": "Final classification: endogenous / exogenous / unverified"
    },
    "Classification Sources": {
        "columns": ["hmdb_origin", "hmdb_origin_category", "coconut_organisms", "chebi_roles"],
        "color": "FFD9B3",
        "description": "Source data used to determine classification"
    },
    "Classification Metadata": {
        "columns": ["classification_basis"],
        "color": "E4DFEC",
        "description": "Reasoning behind classification"
    },
    "DB Support": {
        "columns": ["db_support_level", "db_support_evidence", "mmmdb_detected", "mmmdb_tissues"],
        "color": "FCE4D6",
        "description": "Structure-consensus support from independent databases (not spectral MSI) and MMMDB mouse-tissue evidence"
    },
    "Classification Conflicts": {
        "columns": ["conflict_flag", "conflicting_sources"],
        "color": "F8CBAD",
        "description": "Whether sources disagree on origin, and which verdicts conflict"
    },
    "Enzyme Information": {
        "columns": ["kegg_enzymes", "hmdb_enzymes", "reactome_catalysts", "brenda_enzymes"],
        "color": "DAEEF3",
        "description": "Enzyme data from multiple databases"
    }
}

COL_SOURCE = {
    "InChIKey"             : "COCONUT",
    "compound_name"        : "COCONUT",
    "COCONUT"              : "COCONUT",
    "datasets"             : "pipeline",
    "SMILES"               : "COCONUT",
    "PubChem"              : "PubChem",
    "KEGG"                 : "KEGG",
    "HMDB"                 : "HMDB",
    "ChEBI"                : "ChEBI",
    "CAS"                  : "contributor dataset + HMDB",
    "DrugBank"             : "UniChem",
    "DrugCentral"          : "UniChem",
    "FooDB"                : "UniChem",
    "LIPID MAPS"           : "UniChem",
    "drug_food"            : "DrugBank + DrugCentral + FooDB",
    "classification"       : "ChEBI + HMDB + COCONUT + MMMDB",
    "hmdb_origin"          : "HMDB",
    "hmdb_origin_category" : "HMDB",
    "coconut_organisms"    : "COCONUT",
    "chebi_roles"          : "ChEBI",
    "classification_basis" : "ChEBI + COCONUT + MMMDB",
    "db_support_level"     : "pipeline",
    "db_support_evidence"  : "pipeline",
    "mmmdb_detected"       : "MMMDB",
    "mmmdb_tissues"        : "MMMDB",
    "conflict_flag"        : "ChEBI + HMDB + COCONUT",
    "conflicting_sources"  : "ChEBI + HMDB + COCONUT",
    "kegg_enzymes"         : "KEGG",
    "hmdb_enzymes"         : "HMDB",
    "reactome_catalysts"   : "Reactome",
    "brenda_enzymes"       : "BRENDA", 
}

# External DB IDs 그룹의 헤더(1행)에 걸 각 데이터베이스 공식 홈페이지 링크.
# 클릭하면 해당 사이트로 이동한다(값 셀이 아니라 컬럼 타이틀에만 링크).
DB_HOMEPAGE = {
    "COCONUT"   : "https://coconut.naturalproducts.net/",
    "PubChem"   : "https://pubchem.ncbi.nlm.nih.gov/",
    "KEGG"      : "https://www.kegg.jp/",
    "HMDB"      : "https://hmdb.ca/",
    "ChEBI"     : "https://www.ebi.ac.uk/chebi/",
    "CAS"       : "https://commonchemistry.cas.org/",
    "DrugBank"  : "https://go.drugbank.com/",
    "DrugCentral": "https://drugcentral.org/",
    "FooDB"     : "https://foodb.ca/",
    "LIPID MAPS": "https://www.lipidmaps.org/",
}

COL_DESC = {
    "InChIKey"             : "Chemical structure key (27-char) — primary key for all joins/merges/sorting",
    "compound_name"        : "Compound name",
    "COCONUT"              : "COCONUT compound ID(s) (CNP) mapped to this InChIKey (display reference; semicolon-separated)",
    "datasets"             : "Source datasets this compound appears in (human / mouse_serum / mouse_feces; semicolon-separated)",
    "SMILES"               : "Chemical structure in text format",
    "PubChem"              : "PubChem Compound ID (CID)",
    "KEGG"                 : "KEGG Compound ID",
    "HMDB"                 : "Human Metabolome Database ID",
    "ChEBI"                : "Chemical Entities of Biological Interest ID",
    "CAS"                  : "CAS Registry Number. Not resolved via UniChem — supplied by contributor datasets (osaka(1)) or the HMDB local record; see the Legend Source column",
    "DrugBank"             : "DrugBank ID (cross-linked via UniChem)",
    "DrugCentral"          : "DrugCentral ID (approved-drug DB, cross-linked via UniChem). One of the drug_food evidence sources (drug = DrugBank/DrugCentral present)",
    "FooDB"                : "FooDB ID (cross-linked via UniChem)",
    "LIPID MAPS"           : "LIPID MAPS ID (cross-linked via UniChem)",
    "drug_food"            : "External drug/food signal (drug = DrugBank/DrugCentral present, food = FooDB present). Display-only flag placed before classification; rows are NOT filtered out. Note: FooDB presence is a detection axis and includes many endogenous compounds",
    "classification"       : "Final classification: endogenous / exogenous / unverified",
    "hmdb_origin"          : "HMDB origin field (Endogenous / Food / Drug etc.)",
    "hmdb_origin_category" : "HMDB origin rolled up into 6 buckets (species names dropped) — the origin 'type' at a glance",
    "coconut_organisms"    : "Organisms associated with compound in COCONUT (used to infer classification)",
    "chebi_roles"          : "Role classification from ChEBI (semicolon-separated)",
    "classification_basis" : "Reason behind classification (e.g. MMMDB: detected in mouse tissue / ChEBI: human metabolite)",
    "db_support_level"     : "DB-support level (structure-consensus proxy, NOT spectral MSI): L2 = >=2 independent DB IDs / L3 = <=1 / L4 formula-only / L5 unknown",
    "db_support_evidence"  : "Independent databases supporting this structure (InChIKey + cross-references, incl. MMMDB tissue evidence)",
    "mmmdb_detected"       : "True if compound detected in real mouse tissue (MMMDB, InChIKey match)",
    "mmmdb_tissues"        : "Mouse tissues where compound was detected in MMMDB (semicolon-separated)",
    "conflict_flag"        : "True if sources disagree on endogenous vs exogenous origin",
    "conflicting_sources"  : "Per-source verdicts when they conflict (e.g. COCONUT=endogenous;ChEBI=exogenous)",
    "kegg_enzymes"         : "EC numbers from KEGG (semicolon-separated)",
    "hmdb_enzymes"         : "Enzyme gene names from HMDB (semicolon-separated)",
    "reactome_catalysts"   : "Catalyst activity names from Reactome",
    "brenda_enzymes"       : "EC numbers from BRENDA (semicolon-separated; source: BRENDA Enzyme Database)",
}

def check_column_registration(df_columns):
    """export되는 실제 컬럼과 이 파일의 등록 딕셔너리(GROUPS/COL_SOURCE/COL_DESC)를
    대조해, '한 곳만 바꾸고 나머지를 빠뜨린' 종류의 불일치를 자동으로 잡는다.

    이 검사가 있으면 사람이 눈으로 컬럼을 하나하나 확인할 필요가 없다.
    과거에 이걸로 걸렸어야 했던 것:
      - hmdb_origin_category: GROUPS 미등록 → extras로 밀려 맨 끝 컬럼(AD)에 배치
      - DrugCentral: 근거로 인용되는데 export 컬럼이 없어 감사 불가

    반환: 경고 문자열 리스트(빈 리스트면 이상 없음).
    """
    cols = set(df_columns)
    grouped = {c for g in GROUPS.values() for c in g["columns"]}
    warns = []

    # (1) GROUPS에 없어서 맨 끝 'extras'로 밀리는 컬럼
    unplaced = [c for c in df_columns if c not in grouped]
    if unplaced:
        warns.append(f"[배치] GROUPS 미등록 → 맨 끝으로 밀림: {unplaced} "
                     f"(format_excel.py GROUPS의 적절한 그룹에 추가하세요)")

    # (2) 표시되는 컬럼인데 소스/설명 메타가 없음 → Legend에서 비어 보임
    no_source = [c for c in df_columns if c not in COL_SOURCE]
    no_desc = [c for c in df_columns if c not in COL_DESC]
    if no_source:
        warns.append(f"[Legend] COL_SOURCE 미등록(출처 칸 빔): {no_source}")
    if no_desc:
        warns.append(f"[Legend] COL_DESC 미등록(설명 칸 빔): {no_desc}")

    # (3) drug/food 판정 소스로 쓰이는 DB인데 그 ID 컬럼이 export에 없음 → 근거 감사 불가
    #     (normalize.py의 DRUG_SITES/FOOD_SITES와 동일 집합; 근거는 컬럼으로 노출돼야 함)
    evidence_dbs = {"DrugBank", "DrugCentral", "FooDB"}
    missing_evidence = sorted(evidence_dbs - cols)
    if "drug_food" in cols and missing_evidence:
        warns.append(f"[감사] drug_food 판정 근거 DB인데 컬럼이 없음: {missing_evidence} "
                     f"(근거를 감사할 수 없음 → export_view.py에 컬럼 추가)")

    return warns


def apply_format(filepath: str):
    import pandas as pd
    df = pd.read_excel(filepath)

    # 컬럼 등록 일관성 자동 검사 (한 곳 바꾸고 딸린 곳 빠뜨린 것을 stderr로 경고)
    _warns = check_column_registration(list(df.columns))
    if _warns:
        import sys
        print("⚠ format_excel 컬럼 일관성 경고:", file=sys.stderr)
        for _w in _warns:
            print("   " + _w, file=sys.stderr)

    # GROUPS 정의 순서대로 컬럼 재배치 (Legend 계층과 물리 순서 일치)
    ordered = [c for g in GROUPS.values() for c in g["columns"] if c in df.columns]
    extras = [c for c in df.columns if c not in ordered]
    df = df[ordered + extras]
    df.to_excel(filepath, index=False)

    col_color = {}
    for group, info in GROUPS.items():
        for col in info["columns"]:
            col_color[col] = info["color"]

    wb = load_workbook(filepath)
    ws = wb.active

    # 헤더 볼드 + 배경색 (External DB IDs는 헤더에 홈페이지 하이퍼링크)
    for cell in ws[1]:
        col_name = cell.value
        color = col_color.get(col_name, "FFFFFF")
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        url = DB_HOMEPAGE.get(col_name)
        if url:
            cell.hyperlink = url
            # 링크가 걸린 헤더: 밑줄 + 진한 파랑으로 클릭 가능함을 표시
            cell.font = Font(bold=True, underline="single", color="0563C1")
        else:
            cell.font = Font(bold=True)

    # 컬럼 너비 자동 조정
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # ─────────────────────────────
    # Legend 시트
    # ─────────────────────────────
    if "Legend" in wb.sheetnames:
        del wb["Legend"]
    legend_ws = wb.create_sheet("Legend")

    legend_ws["A1"] = "Metabolite Database — Column Legend"
    legend_ws["A1"].font = Font(bold=True, size=13)
    legend_ws.merge_cells("A1:D1")

    for i, h in enumerate(["Group", "Color", "Column", "Description", "Source"], 1):
        cell = legend_ws.cell(row=3, column=i, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    row = 4
    for group, info in GROUPS.items():
        for col in info["columns"]:
            legend_ws.cell(row=row, column=1, value=group)
            legend_ws.cell(row=row, column=2).fill = PatternFill(
                start_color=info["color"], end_color=info["color"], fill_type="solid"
            )
            legend_ws.cell(row=row, column=3, value=col)
            legend_ws.cell(row=row, column=4, value=COL_DESC.get(col, ""))
            legend_ws.cell(row=row, column=5, value=COL_SOURCE.get(col, ""))
            row += 1

    legend_ws.column_dimensions["A"].width = 22
    legend_ws.column_dimensions["B"].width = 10
    legend_ws.column_dimensions["C"].width = 25
    legend_ws.column_dimensions["D"].width = 50
    legend_ws.column_dimensions["E"].width = 20

    # ─────────────────────────────
    # Summary 시트
    # ─────────────────────────────
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    summary_ws = wb.create_sheet("Summary")

    summary_ws["A1"] = "Metabolite Database — Coverage Summary"
    summary_ws["A1"].font = Font(bold=True, size=13)
    summary_ws.merge_cells("A1:D1")

    summary_ws["A2"] = f"Last updated: {date.today()}"
    summary_ws["A2"].font = Font(italic=True)

    # 헤더
    for i, h in enumerate(["Category", "Item", "Count", "Coverage"], 1):
        cell = summary_ws.cell(row=4, column=i, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    total = len(df)

    def pct(n):
        """비율 문자열. total=0(해당 데이터셋 캐시 없음)일 때 0으로 나누지 않는다."""
        return f"{n / total * 100:.1f}%" if total else "-"

     # 효소 플래그
    def hv(col):
        return df[col].notna() & (df[col].astype(str).str.strip() != "")

    has_kegg     = hv('kegg_enzymes')
    has_hmdb     = hv('hmdb_enzymes')
    has_reactome = hv('reactome_catalysts')
    has_brenda   = hv('brenda_enzymes')

    # v4 classification 나열 문자열(소문자) — endo/exo 근거 포함 여부 카운트에 사용.
    _cls = df['classification'].fillna('').astype(str).str.lower()

    rows = [
        ("Overview",       "Total compounds", total,                        "100%"),
        ("Overview",       "InChIKey",        df['InChIKey'].notna().sum(), pct(df['InChIKey'].notna().sum())),
        ("Overview",       "SMILES",          df['SMILES'].notna().sum(),   pct(df['SMILES'].notna().sum())),
        ("External DB IDs","PubChem",         df['PubChem'].notna().sum(),  pct(df['PubChem'].notna().sum())),
        ("External DB IDs","KEGG",            df['KEGG'].notna().sum(),     pct(df['KEGG'].notna().sum())),
        ("External DB IDs","HMDB",            df['HMDB'].notna().sum(),     pct(df['HMDB'].notna().sum())),
        ("External DB IDs","ChEBI",           df['ChEBI'].notna().sum(),    pct(df['ChEBI'].notna().sum())),
        # v4: classification은 "ChEBI:endogenous; HMDB:exogenous" 나열 문자열.
        # 한 화합물에 endo·exo 근거가 동시에 있을 수 있어 배타 카운트가 아니라
        # '해당 근거를 포함하는 화합물 수'로 센다(합계가 total을 넘을 수 있음).
        ("Classification", "Has endogenous evidence", int(_cls.str.contains('endogenous').sum()), pct(int(_cls.str.contains('endogenous').sum()))),
        ("Classification", "Has exogenous evidence",  int(_cls.str.contains('exogenous').sum()),  pct(int(_cls.str.contains('exogenous').sum()))),
        ("Classification", "Unverified (no evidence)", int((_cls=='unverified').sum()), pct(int((_cls=='unverified').sum()))),
        ("Classification", "Conflict (endo & exo)",   int(df['conflict_flag'].astype(str).str.lower().isin(['true','1']).sum()) if 'conflict_flag' in df.columns else 0, ""),
    ]

    if 'origin_evidence' in df.columns:
        ev = df['origin_evidence'].astype(str)
        def cnt(key): return int(ev.str.contains(key, case=False, na=False).sum())
        rows += [
            ("Origin Evidence", "Reclassified → endogenous", cnt('Homo sapiens'), pct(cnt('Homo sapiens'))),
            ("Origin Evidence", "Reclassified → exogenous", cnt('non-human'), pct(cnt('non-human'))),
            ("Origin Evidence", "Unverified: no organism data", cnt('no organism'), pct(cnt('no organism'))),
            ("Origin Evidence", "Unverified: not in COCONUT", cnt('not in release'), pct(cnt('not in release'))),
        ]

    rows += [
            # 단독
            ("Enzyme Info (single source)", "KEGG only",     int((has_kegg  & ~has_hmdb & ~has_reactome & ~has_brenda).sum()), pct((has_kegg & ~has_hmdb & ~has_reactome & ~has_brenda).sum())),
            ("Enzyme Info (single source)", "HMDB only",     int((~has_kegg & has_hmdb  & ~has_reactome & ~has_brenda).sum()), pct((~has_kegg & has_hmdb & ~has_reactome & ~has_brenda).sum())),
            ("Enzyme Info (single source)", "Reactome only", int((~has_kegg & ~has_hmdb & has_reactome  & ~has_brenda).sum()), pct((~has_kegg & ~has_hmdb & has_reactome & ~has_brenda).sum())),
            ("Enzyme Info (single source)", "BRENDA only",   int((~has_kegg & ~has_hmdb & ~has_reactome & has_brenda).sum()),  pct((~has_kegg & ~has_hmdb & ~has_reactome & has_brenda).sum())),
            # 중복
            ("Enzyme Info (overlap)", "KEGG + HMDB",              int((has_kegg & has_hmdb  & ~has_reactome & ~has_brenda).sum()), pct((has_kegg & has_hmdb & ~has_reactome & ~has_brenda).sum())),
            ("Enzyme Info (overlap)", "KEGG + Reactome",          int((has_kegg & ~has_hmdb & has_reactome  & ~has_brenda).sum()), pct((has_kegg & ~has_hmdb & has_reactome & ~has_brenda).sum())),
            ("Enzyme Info (overlap)", "KEGG + BRENDA",            int((has_kegg & ~has_hmdb & ~has_reactome & has_brenda).sum()),  pct((has_kegg & ~has_hmdb & ~has_reactome & has_brenda).sum())),
            ("Enzyme Info (overlap)", "HMDB + Reactome",          int((~has_kegg & has_hmdb & has_reactome  & ~has_brenda).sum()), pct((~has_kegg & has_hmdb & has_reactome & ~has_brenda).sum())),
            ("Enzyme Info (overlap)", "HMDB + BRENDA",            int((~has_kegg & has_hmdb & ~has_reactome & has_brenda).sum()),  pct((~has_kegg & has_hmdb & ~has_reactome & has_brenda).sum())),
            ("Enzyme Info (overlap)", "Reactome + BRENDA",        int((~has_kegg & ~has_hmdb & has_reactome & has_brenda).sum()),  pct((~has_kegg & ~has_hmdb & has_reactome & has_brenda).sum())),
            ("Enzyme Info (overlap)", "KEGG + HMDB + BRENDA",     int((has_kegg & has_hmdb & ~has_reactome & has_brenda).sum()),   pct((has_kegg & has_hmdb & ~has_reactome & has_brenda).sum())),
            ("Enzyme Info (overlap)", "KEGG + Reactome + BRENDA", int((has_kegg & ~has_hmdb & has_reactome & has_brenda).sum()),   pct((has_kegg & ~has_hmdb & has_reactome & has_brenda).sum())),
            ("Enzyme Info (overlap)", "HMDB + Reactome + BRENDA", int((~has_kegg & has_hmdb & has_reactome & has_brenda).sum()),   pct((~has_kegg & has_hmdb & has_reactome & has_brenda).sum())),
            ("Enzyme Info (overlap)", "All 4 sources",            int((has_kegg & has_hmdb & has_reactome & has_brenda).sum()),    pct((has_kegg & has_hmdb & has_reactome & has_brenda).sum())),
            # 합계
            ("Enzyme Info", "Total (unique)", int((has_kegg | has_hmdb | has_reactome | has_brenda).sum()), pct((has_kegg | has_hmdb | has_reactome | has_brenda).sum())),
        ]

    for i, (category, item, count, coverage) in enumerate(rows, 5):
        summary_ws.cell(row=i, column=1, value=category)
        summary_ws.cell(row=i, column=2, value=item)
        summary_ws.cell(row=i, column=3, value=count)
        summary_ws.cell(row=i, column=4, value=coverage)

    summary_ws.column_dimensions["A"].width = 20
    summary_ws.column_dimensions["B"].width = 25
    summary_ws.column_dimensions["C"].width = 10
    summary_ws.column_dimensions["D"].width = 12

    _build_classification_rules_sheet(wb)

    wb.save(filepath)
    print(f"포맷 적용 완료: {filepath}")


def _build_classification_rules_sheet(wb):
    """
    classification(v4)이 어떻게 만들어지는지 명시하는 시트.
    영어 블록 먼저, 이어서 한국어 블록. 내용은 pipeline/classify.py의
    classify_row_v4() 실제 로직과 1:1 대응(추측 아님).

    v4(2026-07-27 회의): 우선순위 규칙 폐기. 각 DB의 판정을 그대로 나열한다.
    """
    ws = wb.create_sheet("Classification Rules")

    H1  = Font(bold=True, size=14)
    H2  = Font(bold=True, size=12, color="1F4E79")
    BOLD = Font(bold=True)
    ITAL = Font(italic=True, color="595959")
    ENDO_FILL = PatternFill("solid", fgColor="C6EFCE")   # green
    EXO_FILL  = PatternFill("solid", fgColor="FFD9B3")   # orange
    UNV_FILL  = PatternFill("solid", fgColor="D9D9D9")   # grey
    HEAD_FILL = PatternFill("solid", fgColor="BDD7EE")   # blue

    r = 1
    def line(text, font=None, fill=None, col=1):
        nonlocal r
        c = ws.cell(row=r, column=col, value=text)
        if font: c.font = font
        if fill: c.fill = fill
        r += 1
        return c

    def rule_row(code, condition, verdict, fill):
        nonlocal r
        vals = [code, condition, verdict]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.fill = fill
            if j == 1:
                c.font = BOLD
        r += 1

    # ================= ENGLISH =================
    line("Classification method (v4) — each DB's verdict, in parallel", H1); r += 1
    line("The `classification` column does NOT pick one winner. It lists each database's "
         "own verdict side by side, because each DB measures a different axis "
         "(source: pipeline/classify.py → classify_row_v4). Set by the 2026-07-27 advisor meeting.", ITAL); r += 1

    line("What signal each database reads (its axis):", H2)
    for j, h in enumerate(["Database", "Endogenous signal → / Exogenous signal", "Axis"], start=1):
        c = ws.cell(row=r, column=j, value=h); c.font = BOLD; c.fill = HEAD_FILL
    r += 1
    rule_row("ChEBI",   "role 'human metabolite' → endo / any other role → exo", "produced by", ENDO_FILL)
    rule_row("HMDB",    "source 'Endogenous' → endo / Food/Drug/Microbial/Plant/Toxin → exo", "detected in", EXO_FILL)
    rule_row("COCONUT", "organisms incl. Homo sapiens → endo / non-human only → exo", "isolated from", EXO_FILL)
    rule_row("MMMDB",   "detected in real mouse tissue (>=1) → endo (endogenous only)", "measured in tissue", ENDO_FILL)
    r += 1

    line("Output form: `classification` = 'ChEBI:exogenous; HMDB:endogenous; COCONUT:endogenous' "
         "— every DB that has a verdict is listed (order ChEBI>HMDB>COCONUT>MMMDB is for reading "
         "only, NOT a priority). If no DB can decide → 'unverified'.", None); r += 1

    line("Why no single label? (the project's core principle):", H2)
    line("  1. Each DB answers a different question (produced / detected / isolated-from). "
         "Forcing one endo-or-exo answer collapses those distinct axes into a false consensus.", None)
    line("  2. So we never pick a winner — we show all the evidence and let the reader judge. "
         "`conflict_flag` + `conflicting_sources` still surface endo↔exo disagreement for audit; "
         "they no longer drive a merge.", None); r += 1

    line("Rule history is stacked, not overwritten: v1–v3 (priority-based single label) are kept "
         "frozen in classify.py for back-compat and before/after comparison. v4 is the current output.", ITAL); r += 1

    line("conflict_flag = True  when at least one source votes endogenous AND at least "
         "one votes exogenous (MMMDB tissue detection counts as an endogenous vote).", None)
    line("conflicting_sources  lists each source's independent verdict, sorted, e.g. "
         "'ChEBI=endogenous;COCONUT=exogenous;HMDB=exogenous'.", None); r += 2

    # ================= KOREAN =================
    line("분류 방식 (v4) — 각 DB의 판정을 그대로 나열", H1); r += 1
    line("`classification` 컬럼은 하나의 정답을 고르지 않는다. 각 DB가 서로 다른 축을 재기 "
         "때문에, 각 DB의 판정을 나란히 나열한다 "
         "(출처: pipeline/classify.py → classify_row_v4). 2026-07-27 교수님 회의 결정.", ITAL); r += 1

    line("각 DB가 읽는 신호(그 DB의 축):", H2)
    for j, h in enumerate(["데이터베이스", "내인성 신호 / 외인성 신호", "축"], start=1):
        c = ws.cell(row=r, column=j, value=h); c.font = BOLD; c.fill = HEAD_FILL
    r += 1
    rule_row("ChEBI",   "role에 'human metabolite' → endo / 다른 role → exo", "누가 만들었나", ENDO_FILL)
    rule_row("HMDB",    "source가 'Endogenous' → endo / Food/Drug/Microbial/Plant/Toxin → exo", "어디서 검출됐나", EXO_FILL)
    rule_row("COCONUT", "organisms에 Homo sapiens 포함 → endo / 비인간만 → exo", "무엇에서 분리됐나", EXO_FILL)
    rule_row("MMMDB",   "실제 쥐 조직에 검출(1개 이상) → endo (내인성 전용)", "조직에서 실측", ENDO_FILL)
    r += 1

    line("출력 형태: `classification` = 'ChEBI:exogenous; HMDB:endogenous; COCONUT:endogenous' "
         "— 판정이 있는 DB를 모두 나열한다(ChEBI>HMDB>COCONUT>MMMDB 순서는 가독성용일 뿐 "
         "우선순위 아님). 아무 DB도 판정 못 하면 → 'unverified'.", None); r += 1

    line("왜 단일 라벨을 안 만드나? (이 프로젝트의 핵심 원칙):", H2)
    line("  1. 각 DB는 서로 다른 질문(만들었나 / 검출됐나 / 분리됐나)에 답한다. "
         "endo·exo 중 하나로 강제하면 이 서로 다른 축을 뭉개 가짜 합의를 만든다.", None)
    line("  2. 그래서 정답을 고르지 않는다 — 근거를 다 보여주고 읽는 사람이 판단하게 한다. "
         "`conflict_flag`·`conflicting_sources`는 endo↔exo 불일치를 감사용으로 여전히 노출하지만, "
         "더 이상 병합을 강제하지 않는다.", None); r += 1

    line("규칙 이력은 덮어쓰지 않고 쌓는다: v1~v3(우선순위 기반 단일 라벨)는 하위호환·전후 비교를 위해 "
         "classify.py에 박제로 남아있다. v4가 현재 출력이다.", ITAL); r += 1

    line("conflict_flag = True  : endogenous 근거와 exogenous 근거가 동시에 존재할 때 "
         "(MMMDB 조직 검출은 endogenous 표로 계산).", None)
    line("conflicting_sources   : 각 소스의 독립 판정을 정렬해 나열, 예 "
         "'ChEBI=endogenous;COCONUT=exogenous;HMDB=exogenous'.", None)

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 78
    ws.column_dimensions["C"].width = 14