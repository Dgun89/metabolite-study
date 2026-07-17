"""Write mouse_final_curated.xlsx with 3 styled sheets (step30 layout)."""
import pandas as pd
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

sheet1=pd.read_parquet("_sheet1.parquet")
legend=pd.read_pickle("_legend.pkl")
m=pd.read_parquet("mouse_msi.parquet")
N=len(m)
def pct(x): return f"{100*x/N:.1f}%"

# ---- Summary rows ----
orig=m["classification_original"].value_counts()
mm=m["classification"].value_counts()
msi=m["msi_level"].value_counts()
nrec=(m["mmmdb_reclassified"]=="Yes").sum()
nmatch=(m["mmmdb_match"]=="Yes").sum()
exo2endo=((m.classification_original=="exogenous")&(m.classification=="endogenous")).sum()
unv2endo=((m.classification_original=="unverified")&(m.classification=="endogenous")).sum()

S=[["Metabolite Database — Coverage Summary (MMMDB + MSI curated)","","",""],
   [f"Last updated: {date.today().isoformat()}","","",""],
   ["","","",""],
   ["Category","Item","Count","Coverage"],
   ["Overview","Total compounds",N,"100%"],
   ["Overview","InChIKey",int(m.InChIKey.notna().sum()),pct(m.InChIKey.notna().sum())],
   ["Overview","SMILES",int(m.SMILES.notna().sum()),pct(m.SMILES.notna().sum())],
   ["External DB IDs","PubChem",int(m.PubChem.notna().sum()),pct(m.PubChem.notna().sum())],
   ["External DB IDs","KEGG",int(m.KEGG.notna().sum()),pct(m.KEGG.notna().sum())],
   ["External DB IDs","HMDB",int(m.HMDB.notna().sum()),pct(m.HMDB.notna().sum())],
   ["External DB IDs","ChEBI",int(m.ChEBI.notna().sum()),pct(m.ChEBI.notna().sum())],
   ["Classification (MMMDB-aware)","Endogenous",int(mm.get("endogenous",0)),pct(mm.get("endogenous",0))],
   ["Classification (MMMDB-aware)","Exogenous",int(mm.get("exogenous",0)),pct(mm.get("exogenous",0))],
   ["Classification (MMMDB-aware)","Unverified",int(mm.get("unverified",0)),pct(mm.get("unverified",0))],
   ["Classification (original)","Endogenous (before MMMDB)",int(orig.get("endogenous",0)),pct(orig.get("endogenous",0))],
   ["Classification (original)","Exogenous (before MMMDB)",int(orig.get("exogenous",0)),pct(orig.get("exogenous",0))],
   ["Classification (original)","Unverified (before MMMDB)",int(orig.get("unverified",0)),pct(orig.get("unverified",0))],
   ["MMMDB Bridge","Matched in MMMDB (mouse tissues)",int(nmatch),pct(nmatch)],
   ["MMMDB Bridge","Reclassified to endogenous",int(nrec),pct(nrec)],
   ["MMMDB Bridge","  exogenous -> endogenous",int(exo2endo),pct(exo2endo)],
   ["MMMDB Bridge","  unverified -> endogenous",int(unv2endo),pct(unv2endo)],
   ["MSI Confidence","L2 probable (>=2 independent DB)",int(msi.get("L2",0)),pct(msi.get("L2",0))],
   ["MSI Confidence","L3 tentative (structure, <=1 DB)",int(msi.get("L3",0)),pct(msi.get("L3",0))],
   ["MSI Confidence","L4 molecular formula only",int(msi.get("L4",0)),pct(msi.get("L4",0))],
   ["MSI Confidence","L5 unknown (no evidence)",int(msi.get("L5",0)),pct(msi.get("L5",0))],
  ]
summary=pd.DataFrame(S)

# ---- styling ----
HEAD=PatternFill("solid",fgColor="1F4E78"); HEADF=Font(color="FFFFFF",bold=True)
GRP={"Basic Identifiers":"DDEBF7","External DB IDs":"E2EFDA","Classification":"FCE4D6",
     "Classification Sources":"FFF2CC","Classification Metadata":"FFF2CC",
     "Enzyme Information":"EDEDED","MMMDB Bridge":"D9E1F2","MSI Confidence":"E1D5E7",
     "Note":"F2F2F2"}
thin=Side(style="thin",color="BFBFBF"); BORD=Border(thin,thin,thin,thin)

wb=Workbook(); ws=wb.active; ws.title="Sheet1"
# Sheet1
ws.append(list(sheet1.columns))
for j in range(1,len(sheet1.columns)+1):
    c=ws.cell(1,j); c.fill=HEAD; c.font=HEADF; c.alignment=Alignment(horizontal="center"); c.border=BORD
for r in dataframe_to_rows(sheet1,index=False,header=False):
    ws.append(["" if pd.isna(x) else x for x in r])
ws.freeze_panes="A2"
# highlight MSI + reclassified cols
cols=list(sheet1.columns)
msi_col=cols.index("msi_level")+1; rec_col=cols.index("mmmdb_reclassified")+1; mm_col=cols.index("mmmdb_match")+1
MSIF={"L2":"C6EFCE","L3":"FFEB9C","L4":"FFD9B3","L5":"F2F2F2"}
for i in range(2,N+2):
    v=ws.cell(i,msi_col).value
    if v in MSIF: ws.cell(i,msi_col).fill=PatternFill("solid",fgColor=MSIF[v])
    if ws.cell(i,rec_col).value=="Yes":
        ws.cell(i,rec_col).fill=PatternFill("solid",fgColor="C6EFCE")
    if ws.cell(i,mm_col).value=="Yes":
        ws.cell(i,mm_col).fill=PatternFill("solid",fgColor="D9E1F2")
for j,cn in enumerate(cols,1):
    w=min(max(12,len(str(cn))+2),40); ws.column_dimensions[get_column_letter(j)].width=w

# Legend
wl=wb.create_sheet("Legend")
wl["A1"]="Metabolite Database — Column Legend"; wl["A1"].font=Font(bold=True,size=13)
start=3
for ri,row in legend.iterrows():
    rr=start+ri
    for cj in range(5):
        val=row[cj] if cj < len(row) else ""
        cell=wl.cell(rr,cj+1,val)
        cell.border=BORD; cell.alignment=Alignment(vertical="top",wrap_text=(cj==3))
    if ri==0:
        for cj in range(5): wl.cell(rr,cj+1).fill=HEAD; wl.cell(rr,cj+1).font=HEADF
    else:
        g=row[0]
        if g in GRP:
            wl.cell(rr,1).fill=PatternFill("solid",fgColor=GRP[g])
widths=[24,8,26,60,22]
for j,w in enumerate(widths,1): wl.column_dimensions[get_column_letter(j)].width=w

# Summary
wsu=wb.create_sheet("Summary")
for ri,row in summary.iterrows():
    for cj in range(4):
        val=row[cj] if cj<len(row) else ""
        cell=wsu.cell(ri+1,cj+1,val)
        if ri==3: cell.fill=HEAD; cell.font=HEADF; cell.border=BORD
        elif ri<2: cell.font=Font(bold=(ri==0),size=13 if ri==0 else 10)
        elif ri>3: cell.border=BORD
wsu["A1"].font=Font(bold=True,size=13)
for j,w in enumerate([30,36,10,12],1): wsu.column_dimensions[get_column_letter(j)].width=w
# color-band summary category groups
CATF={"MMMDB Bridge":"D9E1F2","MSI Confidence":"E1D5E7","Classification (MMMDB-aware)":"FCE4D6"}
for ri in range(4,len(summary)):
    cat=summary.iloc[ri,0]
    if cat in CATF: wsu.cell(ri+1,1).fill=PatternFill("solid",fgColor=CATF[cat])

wb.save("mouse_final_curated.xlsx")
print("wrote mouse_final_curated.xlsx")
print("sheets:", wb.sheetnames)
print("Sheet1:", N, "rows x", len(cols), "cols")
