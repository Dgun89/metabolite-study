import openpyxl
import requests
import time

# cid = "3845"
# url = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug_view/data/compound/{cid}/JSON"

def get_hmdb_kegg(cid):
    url = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug_view/data/compound/{cid}/JSON"
    response = requests.get(url)

    if response.status_code != 200:
        return {"hmdb": None, "kegg": None}
    
    data = response.json()
    result = {"hmdb": None, "kegg": None}

# print(response.status_code)
# print(data.keys())

    sections = data["Record"]["Section"]
    for section in sections:
        if section.get("TOCHeading") == "Names and Identifiers":
            for subsection in section.get("Section", []):
                if subsection.get("TOCHeading") == "Other Identifiers":
                    for subsubsection in subsection.get("Section", []):
                        heading = subsubsection.get("TOCHeading")
                        if heading in ["HMDB ID", "KEGG ID"]:
                            for info in subsubsection.get("Information", []):
                                for val in info.get("Value", {}).get("StringWithMarkup", []):
                                    text = val.get("String")
                                    if heading == "HMDB ID":
                                        result["hmdb"] = text
                                    else:
                                        result["kegg"] = text
    return result

print(get_hmdb_kegg("3845"))

wb = openpyxl.load_workbook("metabolites_step1_PubChem.xlsx")
ws = wb.active

for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True)):
    cid = row[4] # Column E

    if cid is None: # Skip
        continue

    ids = get_hmdb_kegg(str(cid))
    ws.cell(row=i+2, column=3).value = ids["kegg"] # Column C
    ws.cell(row=i+2, column=4).value = ids["hmdb"] # Column D

    print(f"{i+1} {row[1]} → KEGG: {ids['kegg']} HMDB: {ids['hmdb']}")
    time.sleep(0.3)

wb.save("metabolites_step2_KEGG_HMDB.xlsx")
print("Saved successfully!")

# print(type(data))  # 어떤 타입인지
# print(type(data["Record"]))  # Record는 어떤 타입인지
# print(data["Record"].keys())   # Record 안에 뭐가 있는지

# print(data["Record"]["RecordTitle"])
# print(data["Record"]["RecordNumber"])

# print(type(data["Record"]["Section"]))
# print(len(data["Record"]["Section"]))

# sections = data["Record"]["Section"]
# print(sections[0]["TOCHeading"])
# print(sections[1]["TOCHeading"])
# print(sections[2]["TOCHeading"])