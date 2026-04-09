import openpyxl
wb = openpyxl.load_workbook("metabolites_step1_PubChem.xlsx")
ws = wb.active

empty = 0 
filled = 0

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if row[4] is None:
        empty += 1
    else:
        filled += 1

print(f"filled: {filled}")
print(f"empty: {empty}")
print(f"success rate: {filled/(filled+empty)*100:.1f}%")