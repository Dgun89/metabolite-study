import openpyxl
wb = openpyxl.load_workbook("metabolites_step2_KEGG_HMDB.xlsx")
ws = wb.active

full = 0         # PubChem + KEGG + HMDB
kegg_only = 0    # PubChem + KEGG only
hmdb_only = 0    # PubChem + HMDB only
pubchem_only = 0 # PubChem only
unresolved = 0   # nothing mapped

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    kegg    = row[2]  # Column C
    hmdb    = row[3]  # Column D
    pubchem = row[4]  # Column E

    if pubchem is None:
        unresolved += 1
    elif kegg is not None and hmdb is not None:
        full += 1
    elif kegg is not None and hmdb is None:
        kegg_only += 1
    elif kegg is None and hmdb is not None:
        hmdb_only += 1
    else:
        pubchem_only += 1

total = full + kegg_only + hmdb_only + pubchem_only + unresolved

print(f"PubChem + KEGG + HMDB (all 3): {full} ({full/total*100:.1f}%)")
print(f"PubChem + KEGG only:           {kegg_only} ({kegg_only/total*100:.1f}%)")
print(f"PubChem + HMDB only:           {hmdb_only} ({hmdb_only/total*100:.1f}%)")
print(f"PubChem only:                  {pubchem_only} ({pubchem_only/total*100:.1f}%)")
print(f"Unresolved (none):             {unresolved} ({unresolved/total*100:.1f}%)")
print(f"Total:                         {total}")

names = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if row[1]:  # column B
        names.append(row[1])

total_rows = len(names)
unique = len(set(names))
duplicates = total_rows - unique

print(f"\n--- Duplicate Check ---")
print(f"total rows:         {total_rows}")
print(f"unique metabolites: {unique}")
print(f"duplicate rows:     {duplicates}")